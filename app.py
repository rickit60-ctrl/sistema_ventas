from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from database import get_db, init_db as init_database

# Detectar tipo de base de datos
DATABASE_URL = os.environ.get('DATABASE_URL')
USE_POSTGRES = DATABASE_URL is not None

if USE_POSTGRES:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    print("✓ Usando PostgreSQL (Producción)")
else:
    # SQLite removido - usando PostgreSQL
    print("✓ Usando SQLite (Desarrollo)")

# Importar adaptador de base de datos
try:
    from db_adapter import adapt_query, get_placeholder, get_cursor
except ImportError:
    # Fallback si db_adapter no existe
    def adapt_query(q):
        if USE_POSTGRES:
            q = q.replace('INTEGER PRIMARY KEY AUTOINCREMENT', 'SERIAL PRIMARY KEY')
            q = q.replace('REAL', 'DECIMAL(10,2)')
            q = q.replace('TIMESTAMP DEFAULT CURRENT_TIMESTAMP', 'TIMESTAMP DEFAULT NOW()')
            q = q.replace('?', '%s')
        return q
    def get_placeholder():
        return '%s' if USE_POSTGRES else '?'
    def get_cursor(conn):
        return conn.cursor(cursor_factory=RealDictCursor) if USE_POSTGRES else conn.cursor()

app = Flask(__name__)

# Configuración de producción
app.secret_key = os.environ.get('SECRET_KEY', 'tu_clave_secreta_super_segura_cambiala_en_produccion')


# ==================== FUNCIONES AUXILIARES ====================


def login_required(f):
    """Decorador para requerir login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor inicia sesión', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def get_config(clave, default=''):
    """Obtener configuración del sistema"""
    try:
        db = get_db()
        config = db.execute('SELECT valor FROM configuracion WHERE clave = %s', (clave,)).fetchone()
        db.close()
        return config['valor'] if config else default
    except:
        return default

@app.context_processor
def inject_config():
    """Inyectar configuración en todos los templates"""
    return {
        'moneda': get_config('moneda_simbolo', 'RD$'),
        'moneda_codigo': get_config('moneda_codigo', 'DOP'),
        'now': datetime.now
    }

# ==================== RUTAS PRINCIPALES ====================

@app.route('/health')
def health():
    """Healthcheck para Railway"""
    return jsonify({
        'status': 'healthy',
        'database': 'connected' if os.environ.get('DATABASE_URL') else 'not_configured'
    }), 200

@app.route('/')
def index():
    """Página principal"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        db = get_db()
        user = db.execute('SELECT * FROM usuarios WHERE username = %s', (username,)).fetchone()
        db.close()
        
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['nombre'] = user['nombre']
            flash('¡Bienvenido al sistema!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Cerrar sesión"""
    session.clear()
    flash('Sesión cerrada exitosamente', 'info')
    return redirect(url_for('login'))

# ==================== DASHBOARD ====================

@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard principal"""
    db = get_db()
    user_id = session['user_id']
    
    # Fecha actual
    hoy = datetime.now()
    mes_actual = hoy.month
    anio_actual = hoy.year
    
    # Total vendido este mes
    total_vendido = db.execute('''
        SELECT COALESCE(SUM(total_vendido), 0) as total
        FROM ventas
        WHERE TO_CHAR(fecha_venta, 'MM') = ? AND TO_CHAR(fecha_venta, 'YYYY') = ? AND usuario_id = ?
    ''', (f'{mes_actual:02d}', str(anio_actual), user_id)).fetchone()['total']
    
    # Ganancia del mes
    ganancia_mes = db.execute('''
        SELECT COALESCE(SUM(ganancia), 0) as total
        FROM ventas
        WHERE TO_CHAR(fecha_venta, 'MM') = ? AND TO_CHAR(fecha_venta, 'YYYY') = ? AND usuario_id = ?
    ''', (f'{mes_actual:02d}', str(anio_actual), user_id)).fetchone()['total']
    
    # Total pendiente por cobrar
    total_pendiente = db.execute('''
        SELECT COALESCE(SUM(v.total_vendido - COALESCE(pagos.total_pagado, 0)), 0) as pendiente
        FROM ventas v
        LEFT JOIN (
            SELECT venta_id, SUM(monto) as total_pagado
            FROM pagos
            GROUP BY venta_id
        ) pagos ON v.id = pagos.venta_id
        WHERE v.tipo_venta = 'credito' AND v.estado_pago != 'completado' AND v.usuario_id = ?
    ''', (user_id,)).fetchone()['pendiente']
    
    # Diezmo del mes
    diezmo_mes = db.execute('''
        SELECT COALESCE(SUM(diezmo), 0) as total
        FROM ventas
        WHERE TO_CHAR(fecha_venta, 'MM') = ? AND TO_CHAR(fecha_venta, 'YYYY') = ? AND usuario_id = ?
    ''', (f'{mes_actual:02d}', str(anio_actual), user_id)).fetchone()['total']
    
    # Valor del inventario
    valor_inventario = db.execute('''
        SELECT COALESCE(SUM(cantidad * costo_unitario), 0) as valor
        FROM productos
        WHERE usuario_id = %s
    ''', (user_id,)).fetchone()['valor']
    
    # Productos con bajo stock
    productos_bajo_stock = db.execute('''
        SELECT COUNT(*) as total
        FROM productos
        WHERE cantidad > 0 AND cantidad <= stock_minimo AND usuario_id = %s
    ''', (user_id,)).fetchone()['total']
    
    # Productos agotados
    productos_agotados = db.execute('''
        SELECT COUNT(*) as total
        FROM productos
        WHERE cantidad = 0 AND usuario_id = %s
    ''', (user_id,)).fetchone()['total']
    
    # Productos críticos
    productos_criticos = db.execute('''
        SELECT nombre, cantidad, stock_minimo
        FROM productos
        WHERE cantidad <= stock_minimo AND usuario_id = %s
        ORDER BY cantidad ASC
        LIMIT 5
    ''', (user_id,)).fetchall()
    
    # Ventas recientes
    ventas_recientes = db.execute('''
        SELECT v.*, p.nombre as producto_nombre
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        WHERE v.usuario_id = %s
        ORDER BY v.fecha_venta DESC, v.fecha_registro DESC
        LIMIT 5
    ''', (user_id,)).fetchall()
    
    # Nombres de meses
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    db.close()
    return render_template('dashboard.html',
                         total_vendido=total_vendido,
                         ganancia_mes=ganancia_mes,
                         total_pendiente=total_pendiente,
                         diezmo_mes=diezmo_mes,
                         valor_inventario=valor_inventario,
                         productos_bajo_stock=productos_bajo_stock,
                         productos_agotados=productos_agotados,
                         productos_criticos=productos_criticos,
                         ventas_recientes=ventas_recientes,
                         mes_nombre=meses[mes_actual-1],
                         anio=anio_actual)

# ==================== GASTOS ====================

@app.route('/gastos')
@login_required
def gastos():
    """Lista de gastos"""
    db = get_db()
    user_id = session['user_id']
    
    # Obtener mes y año actual
    hoy = datetime.now()
    mes_actual = int(request.args.get('mes', hoy.month))
    anio_actual = int(request.args.get('anio', hoy.year))
    
    # Obtener gastos del mes
    gastos_list = db.execute('''
        SELECT * FROM gastos
        WHERE TO_CHAR(fecha, 'MM') = ? AND TO_CHAR(fecha, 'YYYY') = ? AND usuario_id = ?
        ORDER BY fecha DESC
    ''', (f'{mes_actual:02d}', str(anio_actual), user_id)).fetchall()
    
    # Calcular totales por categoría
    totales_categorias = db.execute('''
        SELECT categoria, SUM(monto) as total
        FROM gastos
        WHERE TO_CHAR(fecha, 'MM') = ? AND TO_CHAR(fecha, 'YYYY') = ? AND usuario_id = ?
        GROUP BY categoria
    ''', (f'{mes_actual:02d}', str(anio_actual), user_id)).fetchall()
    
    # Total mensual
    total_mensual = db.execute('''
        SELECT COALESCE(SUM(monto), 0) as total
        FROM gastos
        WHERE TO_CHAR(fecha, 'MM') = ? AND TO_CHAR(fecha, 'YYYY') = ? AND usuario_id = ?
    ''', (f'{mes_actual:02d}', str(anio_actual), user_id)).fetchone()['total']
    
    # Nombres de meses
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    db.close()
    return render_template('gastos.html',
                         gastos=gastos_list,
                         totales_categorias=totales_categorias,
                         total_mensual=total_mensual,
                         mes_actual=mes_actual,
                         anio_actual=anio_actual,
                         mes_nombre=meses[mes_actual-1])

@app.route('/gastos/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_gasto():
    """Registrar nuevo gasto"""
    if request.method == 'POST':
        db = get_db()
        user_id = session['user_id']
        
        fecha = request.form.get('fecha')
        categoria = request.form.get('categoria')
        descripcion = request.form.get('descripcion', '')
        monto = float(request.form.get('monto'))
        
        # Insertar gasto
        db.execute('''
            INSERT INTO gastos (fecha, categoria, descripcion, monto, usuario_id)
            VALUES (%s, %s, %s, %s, %s)
        ''', (fecha, categoria, descripcion, monto, user_id))
        
        db.commit()
        db.close()
        
        flash('Gasto registrado exitosamente', 'success')
        return redirect(url_for('gastos'))
    
    return render_template('nuevo_gasto.html')

@app.route('/gastos/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_gasto(id):
    """Eliminar gasto"""
    db = get_db()
    user_id = session['user_id']
    
    # Verificar que el gasto pertenece al usuario
    gasto = db.execute('SELECT id FROM gastos WHERE id = %s AND usuario_id = %s', (id, user_id)).fetchone()
    
    if gasto:
        db.execute('DELETE FROM gastos WHERE id = %s', (id,))
        db.commit()
        flash('Gasto eliminado exitosamente', 'success')
    else:
        flash('Gasto no encontrado', 'error')
    
    db.close()
    return redirect(url_for('gastos'))

@app.route('/gastos/exportar', methods=['POST'])
@login_required
def exportar_gastos():
    """Exportar gastos quincenales a Excel"""
    db = get_db()
    user_id = session['user_id']
    
    mes = int(request.form.get('mes'))
    anio = int(request.form.get('anio'))
    quincena = request.form.get('quincena')
    
    # Nombres de meses
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    # Definir rango de días según quincena
    if quincena == 'primera':
        dia_inicio = 1
        dia_fin = 15
        nombre_quincena = '1ra Quincena'
    else:
        dia_inicio = 16
        dia_fin = 31
        nombre_quincena = '2da Quincena'
    
    # Obtener gastos de la quincena
    gastos_list = db.execute('''
        SELECT fecha, categoria, descripcion, monto
        FROM gastos
        WHERE TO_CHAR(fecha, 'MM') = ? AND TO_CHAR(fecha, 'YYYY') = ?
        AND CAST(TO_CHAR(fecha, 'DD') AS INTEGER) BETWEEN ? AND ?
        AND usuario_id = ?
        ORDER BY fecha ASC
    ''', (f'{mes:02d}', str(anio), dia_inicio, dia_fin, user_id)).fetchall()
    
    # Calcular total
    total = sum([g['monto'] for g in gastos_list])
    
    db.close()
    
    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Gastos {nombre_quincena}"
    
    # Configurar columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    
    # Título
    ws.merge_cells('A1:D1')
    titulo = ws['A1']
    titulo.value = f'REPORTE DE GASTOS - {nombre_quincena} de {meses[mes-1]} {anio}'
    titulo.font = Font(size=14, bold=True, color='FFFFFF')
    titulo.fill = PatternFill(start_color='0a6ed1', end_color='0a6ed1', fill_type='solid')
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Headers
    headers = ['Fecha', 'Categoría', 'Descripción', 'Monto']
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    header_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    moneda = get_config('moneda_simbolo', 'RD$')
    row = 4
    for gasto in gastos_list:
        ws.cell(row=row, column=1, value=gasto['fecha']).border = border
        ws.cell(row=row, column=2, value=gasto['categoria']).border = border
        ws.cell(row=row, column=3, value=gasto['descripcion'] or '-').border = border
        ws.cell(row=row, column=4, value=f"{moneda}{gasto['monto']:.2f}").border = border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        row += 1
    
    # Total
    row += 1
    ws.cell(row=row, column=3, value='TOTAL:').font = Font(bold=True, size=12)
    ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
    total_cell = ws.cell(row=row, column=4, value=f"{moneda}{total:.2f}")
    total_cell.font = Font(bold=True, size=12, color='FFFFFF')
    total_cell.fill = PatternFill(start_color='107e3e', end_color='107e3e', fill_type='solid')
    total_cell.alignment = Alignment(horizontal='right')
    total_cell.border = border
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'Gastos_{nombre_quincena.replace(" ", "_")}_{meses[mes-1]}_{anio}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ==================== INVENTARIO ====================

@app.route('/inventario')
@login_required
def inventario():
    """Lista de productos en inventario"""
    db = get_db()
    user_id = session['user_id']
    
    productos = db.execute('''
        SELECT * FROM productos
        WHERE usuario_id = %s
        ORDER BY fecha_registro DESC
    ''', (user_id,)).fetchall()
    
    # Valor total del inventario
    valor_total = db.execute('''
        SELECT COALESCE(SUM(cantidad * costo_unitario), 0) as total
        FROM productos
        WHERE usuario_id = %s
    ''', (user_id,)).fetchone()['total']
    
    db.close()
    return render_template('inventario.html', productos=productos, valor_total=valor_total)

@app.route('/inventario/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_producto():
    """Agregar nuevo producto"""
    if request.method == 'POST':
        db = get_db()
        user_id = session['user_id']
        
        nombre = request.form.get('nombre')
        descripcion = request.form.get('descripcion')
        cantidad = int(request.form.get('cantidad'))
        costo_unitario = float(request.form.get('costo_unitario'))
        precio_venta = float(request.form.get('precio_venta'))
        stock_minimo = int(request.form.get('stock_minimo', 5))
        
        # Determinar estado
        if cantidad == 0:
            estado = 'agotado'
        elif cantidad <= stock_minimo:
            estado = 'bajo'
        else:
            estado = 'disponible'
        
        db.execute('''
            INSERT INTO productos (nombre, descripcion, cantidad, costo_unitario, precio_venta, stock_minimo, estado, usuario_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ''', (nombre, descripcion, cantidad, costo_unitario, precio_venta, stock_minimo, estado, user_id))
        
        db.commit()
        db.close()
        
        flash('Producto agregado exitosamente', 'success')
        return redirect(url_for('inventario'))
    
    return render_template('nuevo_producto.html')

@app.route('/inventario/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_producto(id):
    """Editar producto"""
    db = get_db()
    user_id = session['user_id']
    
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        descripcion = request.form.get('descripcion')
        cantidad = int(request.form.get('cantidad'))
        costo_unitario = float(request.form.get('costo_unitario'))
        precio_venta = float(request.form.get('precio_venta'))
        stock_minimo = int(request.form.get('stock_minimo'))
        
        # Determinar estado
        if cantidad == 0:
            estado = 'agotado'
        elif cantidad <= stock_minimo:
            estado = 'bajo'
        else:
            estado = 'disponible'
        
        db.execute('''
            UPDATE productos
            SET nombre = %s, descripcion = %s, cantidad = %s, costo_unitario = %s, precio_venta = %s, stock_minimo = %s, estado = %s
            WHERE id = %s AND usuario_id = %s
        ''', (nombre, descripcion, cantidad, costo_unitario, precio_venta, stock_minimo, estado, id, user_id))
        
        db.commit()
        db.close()
        
        flash('Producto actualizado exitosamente', 'success')
        return redirect(url_for('inventario'))
    
    producto = db.execute('SELECT * FROM productos WHERE id = %s AND usuario_id = %s', (id, user_id)).fetchone()
    db.close()
    
    if not producto:
        flash('Producto no encontrado', 'error')
        return redirect(url_for('inventario'))
    
    return render_template('editar_producto.html', producto=producto)

@app.route('/inventario/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_producto(id):
    """Eliminar producto"""
    db = get_db()
    user_id = session['user_id']
    
    # Verificar si el producto tiene ventas asociadas
    ventas = db.execute('SELECT COUNT(*) as total FROM ventas WHERE producto_id = %s', (id,)).fetchone()['total']
    
    if ventas > 0:
        flash('No se puede eliminar el producto porque tiene ventas asociadas', 'error')
    else:
        db.execute('DELETE FROM productos WHERE id = %s AND usuario_id = %s', (id, user_id))
        db.commit()
        flash('Producto eliminado exitosamente', 'success')
    
    db.close()
    return redirect(url_for('inventario'))

# ==================== VENTAS ====================

@app.route('/ventas')
@login_required
def ventas():
    """Lista de ventas"""
    db = get_db()
    user_id = session['user_id']
    
    ventas_list = db.execute('''
        SELECT v.*, p.nombre as producto_nombre
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        WHERE v.usuario_id = %s
        ORDER BY v.fecha_venta DESC, v.fecha_registro DESC
    ''', (user_id,)).fetchall()
    
    # Totales
    totales = db.execute('''
        SELECT 
            COALESCE(SUM(total_vendido), 0) as total_vendido,
            COALESCE(SUM(ganancia), 0) as total_ganancia,
            COALESCE(SUM(diezmo), 0) as total_diezmo
        FROM ventas
        WHERE usuario_id = %s
    ''', (user_id,)).fetchone()
    
    db.close()
    return render_template('ventas.html', ventas=ventas_list, totales=totales)

@app.route('/ventas/nueva', methods=['GET', 'POST'])
@login_required
def nueva_venta():
    """Registrar nueva venta"""
    db = get_db()
    user_id = session['user_id']
    
    if request.method == 'POST':
        producto_id = int(request.form.get('producto_id'))
        cliente_nombre = request.form.get('cliente_nombre')
        cliente_telefono = request.form.get('cliente_telefono', '')
        cantidad = int(request.form.get('cantidad'))
        tipo_venta = request.form.get('tipo_venta')
        fecha_venta = request.form.get('fecha_venta')
        
        # Obtener producto
        producto = db.execute('SELECT * FROM productos WHERE id = %s AND usuario_id = %s', (producto_id, user_id)).fetchone()
        
        if not producto:
            flash('Producto no encontrado', 'error')
            db.close()
            return redirect(url_for('nueva_venta'))
        
        # Verificar stock
        if producto['cantidad'] < cantidad:
            flash(f'Stock insuficiente. Disponible: {producto["cantidad"]} unidades', 'error')
            db.close()
            return redirect(url_for('nueva_venta'))
        
        # Calcular valores
        precio_unitario = producto['precio_venta']
        total_vendido = precio_unitario * cantidad
        costo_total = producto['costo_unitario'] * cantidad
        ganancia = total_vendido - costo_total
        diezmo = total_vendido * 0.10
        
        # Determinar estado de pago
        estado_pago = 'completado' if tipo_venta == 'contado' else 'pendiente'
        
        # Insertar venta
        cursor = db.execute('''
            INSERT INTO ventas (producto_id, cliente_nombre, cliente_telefono, cantidad, precio_unitario, 
                               total_vendido, costo_total, ganancia, diezmo, tipo_venta, estado_pago, fecha_venta, usuario_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (producto_id, cliente_nombre, cliente_telefono, cantidad, precio_unitario,
              total_vendido, costo_total, ganancia, diezmo, tipo_venta, estado_pago, fecha_venta, user_id))
        
        venta_id = cursor.lastrowid
        
        # Si es venta al contado, registrar pago automático
        if tipo_venta == 'contado':
            db.execute('''
                INSERT INTO pagos (venta_id, monto, fecha_pago, metodo_pago, notas, usuario_id)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (venta_id, total_vendido, fecha_venta, 'Contado', 'Pago completo al contado', user_id))
        
        # Actualizar inventario
        nueva_cantidad = producto['cantidad'] - cantidad
        if nueva_cantidad == 0:
            nuevo_estado = 'agotado'
        elif nueva_cantidad <= producto['stock_minimo']:
            nuevo_estado = 'bajo'
        else:
            nuevo_estado = 'disponible'
        
        db.execute('''
            UPDATE productos
            SET cantidad = %s, estado = %s
            WHERE id = %s
        ''', (nueva_cantidad, nuevo_estado, producto_id))
        
        # Actualizar o crear diezmo mensual
        mes_venta = int(fecha_venta.split('-')[1])
        anio_venta = int(fecha_venta.split('-')[0])
        
        diezmo_existente = db.execute('''
            SELECT * FROM diezmos_mensuales
            WHERE mes = %s AND anio = %s AND usuario_id = %s
        ''', (mes_venta, anio_venta, user_id)).fetchone()
        
        if diezmo_existente:
            db.execute('''
                UPDATE diezmos_mensuales
                SET total_diezmo = total_diezmo + %s
                WHERE mes = %s AND anio = %s AND usuario_id = %s
            ''', (diezmo, mes_venta, anio_venta, user_id))
        else:
            db.execute('''
                INSERT INTO diezmos_mensuales (mes, anio, total_diezmo, usuario_id)
                VALUES (%s, %s, %s, %s)
            ''', (mes_venta, anio_venta, diezmo, user_id))
        
        db.commit()
        db.close()
        
        flash('Venta registrada exitosamente', 'success')
        return redirect(url_for('ventas'))
    
    # GET - mostrar formulario
    productos = db.execute('''
        SELECT * FROM productos
        WHERE cantidad > 0 AND usuario_id = %s
        ORDER BY nombre
    ''', (user_id,)).fetchall()
    
    db.close()
    return render_template('nueva_venta.html', productos=productos)

# ==================== CUENTAS POR COBRAR ====================

@app.route('/cuentas-por-cobrar')
@login_required
def cuentas_por_cobrar():
    """Cuentas por cobrar"""
    db = get_db()
    user_id = session['user_id']
    
    ventas_credito = db.execute('''
        SELECT v.*, p.nombre as producto_nombre,
               COALESCE(pagos.total_pagado, 0) as total_pagado,
               (v.total_vendido - COALESCE(pagos.total_pagado, 0)) as saldo_pendiente
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        LEFT JOIN (
            SELECT venta_id, SUM(monto) as total_pagado
            FROM pagos
            GROUP BY venta_id
        ) pagos ON v.id = pagos.venta_id
        WHERE v.tipo_venta = 'credito' AND v.estado_pago != 'completado' AND v.usuario_id = ?
        ORDER BY v.fecha_venta DESC
    ''', (user_id,)).fetchall()
    
    # Total por cobrar
    total_por_cobrar = sum([v['saldo_pendiente'] for v in ventas_credito])
    
    db.close()
    return render_template('cuentas_por_cobrar.html', ventas=ventas_credito, total_por_cobrar=total_por_cobrar)

@app.route('/pagos/<int:venta_id>')
@login_required
def ver_pagos(venta_id):
    """Ver pagos de una venta"""
    db = get_db()
    user_id = session['user_id']
    
    venta = db.execute('''
        SELECT v.*, p.nombre as producto_nombre
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        WHERE v.id = %s AND v.usuario_id = %s
    ''', (venta_id, user_id)).fetchone()
    
    if not venta:
        flash('Venta no encontrada', 'error')
        db.close()
        return redirect(url_for('cuentas_por_cobrar'))
    
    # Obtener pagos
    pagos = db.execute('''
        SELECT * FROM pagos
        WHERE venta_id = %s
        ORDER BY fecha_pago DESC
    ''', (venta_id,)).fetchall()
    
    # Total pagado
    total_pagado = sum([p['monto'] for p in pagos])
    saldo_pendiente = venta['total_vendido'] - total_pagado
    
    db.close()
    return render_template('ver_pagos.html', venta=venta, pagos=pagos, 
                         total_pagado=total_pagado, saldo_pendiente=saldo_pendiente)

@app.route('/pagos/registrar/<int:venta_id>', methods=['POST'])
@login_required
def registrar_pago(venta_id):
    """Registrar pago de una venta a crédito"""
    db = get_db()
    user_id = session['user_id']
    
    # Verificar venta
    venta = db.execute('SELECT * FROM ventas WHERE id = %s AND usuario_id = %s', (venta_id, user_id)).fetchone()
    
    if not venta:
        flash('Venta no encontrada', 'error')
        db.close()
        return redirect(url_for('cuentas_por_cobrar'))
    
    # Obtener datos del formulario
    monto = float(request.form.get('monto'))
    fecha_pago = request.form.get('fecha_pago')
    metodo_pago = request.form.get('metodo_pago')
    notas = request.form.get('notas', '')
    
    # Calcular total pagado hasta ahora
    total_pagado = db.execute('''
        SELECT COALESCE(SUM(monto), 0) as total
        FROM pagos
        WHERE venta_id = %s
    ''', (venta_id,)).fetchone()['total']
    
    saldo_pendiente = venta['total_vendido'] - total_pagado
    
    # Validar monto
    if monto > saldo_pendiente:
        flash(f'El monto excede el saldo pendiente ({saldo_pendiente:.2f})', 'error')
        db.close()
        return redirect(url_for('ver_pagos', venta_id=venta_id))
    
    # Registrar pago
    db.execute('''
        INSERT INTO pagos (venta_id, monto, fecha_pago, metodo_pago, notas, usuario_id)
        VALUES (%s, %s, %s, %s, %s, %s)
    ''', (venta_id, monto, fecha_pago, metodo_pago, notas, user_id))
    
    # Actualizar estado de la venta
    nuevo_total_pagado = total_pagado + monto
    if nuevo_total_pagado >= venta['total_vendido']:
        nuevo_estado = 'completado'
    else:
        nuevo_estado = 'parcial'
    
    db.execute('''
        UPDATE ventas
        SET estado_pago = %s
        WHERE id = %s
    ''', (nuevo_estado, venta_id))
    
    db.commit()
    db.close()
    
    flash('Pago registrado exitosamente', 'success')
    return redirect(url_for('ver_pagos', venta_id=venta_id))

# ==================== DIEZMOS ====================

@app.route('/diezmos')
@login_required
def diezmos():
    """Diezmos mensuales"""
    db = get_db()
    user_id = session['user_id']
    
    diezmos_list = db.execute('''
        SELECT * FROM diezmos_mensuales
        WHERE usuario_id = %s
        ORDER BY anio DESC, mes DESC
    ''', (user_id,)).fetchall()
    
    # Resumen
    resumen = db.execute('''
        SELECT 
            COALESCE(SUM(CASE WHEN estado = 'Pendiente' THEN total_diezmo ELSE 0 END), 0) as total_pendiente,
            COALESCE(SUM(CASE WHEN estado = 'Entregado' THEN total_diezmo ELSE 0 END), 0) as total_entregado,
            COALESCE(SUM(total_diezmo), 0) as total_general
        FROM diezmos_mensuales
        WHERE usuario_id = ?
    ''', (user_id,)).fetchone()
    
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    db.close()
    return render_template('diezmos.html', diezmos=diezmos_list, resumen=resumen, meses=meses)

@app.route('/diezmos/marcar/<int:id>', methods=['POST'])
@login_required
def marcar_diezmo(id):
    """Marcar diezmo como entregado o pendiente"""
    db = get_db()
    user_id = session['user_id']
    
    diezmo = db.execute('SELECT * FROM diezmos_mensuales WHERE id = %s AND usuario_id = %s', (id, user_id)).fetchone()
    
    if diezmo:
        nuevo_estado = 'Entregado' if diezmo['estado'] == 'Pendiente' else 'Pendiente'
        fecha_entrega = datetime.now().strftime('%Y-%m-%d %H:%M:%S') if nuevo_estado == 'Entregado' else None
        
        db.execute('''
            UPDATE diezmos_mensuales
            SET estado = %s, fecha_entrega = %s
            WHERE id = %s
        ''', (nuevo_estado, fecha_entrega, id))
        
        db.commit()
        flash(f'Diezmo marcado como {nuevo_estado}', 'success')
    
    db.close()
    return redirect(url_for('diezmos'))

# ==================== REPORTES ====================

@app.route('/reportes')
@login_required
def reportes():
    """Página de reportes"""
    return render_template('reportes.html')

@app.route('/reportes/exportar', methods=['POST'])
@login_required
def exportar_reporte():
    """Exportar reporte mensual a Excel"""
    db = get_db()
    user_id = session['user_id']
    
    mes = int(request.form.get('mes'))
    anio = int(request.form.get('anio'))
    
    # Obtener ventas del mes
    ventas = db.execute('''
        SELECT v.*, p.nombre as producto_nombre
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        WHERE TO_CHAR(v.fecha_venta, 'MM') = ? AND TO_CHAR(v.fecha_venta, 'YYYY') = ? AND v.usuario_id = ?
        ORDER BY v.fecha_venta
    ''', (f'{mes:02d}', str(anio), user_id)).fetchall()
    
    db.close()
    
    # Nombres de meses
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    # Crear Excel directamente
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ventas {meses[mes-1]} {anio}"
    
    # Título
    ws.merge_cells('A1:H1')
    titulo = ws['A1']
    titulo.value = f'REPORTE DE VENTAS - {meses[mes-1]} {anio}'
    titulo.font = Font(size=14, bold=True, color='FFFFFF')
    titulo.fill = PatternFill(start_color='0a6ed1', end_color='0a6ed1', fill_type='solid')
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Headers
    headers = ['Fecha', 'Cliente', 'Producto', 'Cantidad', 'Precio Unit.', 'Total', 'Ganancia', 'Diezmo']
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    header_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    moneda = get_config('moneda_simbolo', 'RD$')
    row = 4
    total_vendido = 0
    total_ganancia = 0
    total_diezmo = 0
    
    for venta in ventas:
        ws.cell(row=row, column=1, value=venta['fecha_venta']).border = border
        ws.cell(row=row, column=2, value=venta['cliente_nombre']).border = border
        ws.cell(row=row, column=3, value=venta['producto_nombre']).border = border
        ws.cell(row=row, column=4, value=venta['cantidad']).border = border
        ws.cell(row=row, column=5, value=f"{moneda}{venta['precio_unitario']:.2f}").border = border
        ws.cell(row=row, column=6, value=f"{moneda}{venta['total_vendido']:.2f}").border = border
        ws.cell(row=row, column=7, value=f"{moneda}{venta['ganancia']:.2f}").border = border
        ws.cell(row=row, column=8, value=f"{moneda}{venta['diezmo']:.2f}").border = border
        
        total_vendido += venta['total_vendido']
        total_ganancia += venta['ganancia']
        total_diezmo += venta['diezmo']
        row += 1
    
    # Totales
    row += 1
    ws.cell(row=row, column=5, value='TOTALES:').font = Font(bold=True, size=12)
    ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=6, value=f"{moneda}{total_vendido:.2f}").font = Font(bold=True, size=12)
    ws.cell(row=row, column=6).fill = PatternFill(start_color='107e3e', end_color='107e3e', fill_type='solid')
    ws.cell(row=row, column=7, value=f"{moneda}{total_ganancia:.2f}").font = Font(bold=True, size=12)
    ws.cell(row=row, column=8, value=f"{moneda}{total_diezmo:.2f}").font = Font(bold=True, size=12)
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'Reporte_{meses[mes-1]}_{anio}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ==================== CONFIGURACIÓN ====================

@app.route('/configuracion', methods=['GET', 'POST'])
@login_required
def configuracion():
    """Configuración del sistema"""
    db = get_db()
    user_id = session['user_id']
    
    if request.method == 'POST':
        moneda_simbolo = request.form.get('moneda_simbolo')
        moneda_codigo = request.form.get('moneda_codigo')
        
        db.execute('''
            UPDATE configuracion SET valor = ?, fecha_modificacion = CURRENT_TIMESTAMP
            WHERE clave = 'moneda_simbolo' AND usuario_id = ?
        ''', (moneda_simbolo, user_id))
        
        db.execute('''
            UPDATE configuracion SET valor = ?, fecha_modificacion = CURRENT_TIMESTAMP
            WHERE clave = 'moneda_codigo' AND usuario_id = ?
        ''', (moneda_codigo, user_id))
        
        db.commit()
        flash('Configuración actualizada exitosamente', 'success')
        return redirect(url_for('configuracion'))
    
    config_actual = {
        'moneda_simbolo': get_config('moneda_simbolo', 'RD$'),
        'moneda_codigo': get_config('moneda_codigo', 'DOP')
    }
    
    db.close()
    return render_template('configuracion.html', config=config_actual)

# ==================== API ====================

@app.route('/api/estadisticas')
@login_required
def api_estadisticas():
    """API para estadísticas del dashboard"""
    db = get_db()
    user_id = session['user_id']
    
    estadisticas = []
    hoy = datetime.now()
    
    for i in range(5, -1, -1):
        fecha = hoy - timedelta(days=30*i)
        mes = fecha.month
        anio = fecha.year
        
        total = db.execute('''
            SELECT COALESCE(SUM(total_vendido), 0) as total
            FROM ventas
            WHERE TO_CHAR(fecha_venta, 'MM') = ? AND TO_CHAR(fecha_venta, 'YYYY') = ? AND usuario_id = ?
        ''', (f'{mes:02d}', str(anio), user_id)).fetchone()['total']
        
        estadisticas.append({
            'mes': fecha.strftime('%b'),
            'total': float(total)
        })
    
    db.close()
    return jsonify(estadisticas)

@app.route('/api/producto/<int:id>')
@login_required
def api_producto(id):
    """API para obtener datos de un producto"""
    db = get_db()
    user_id = session['user_id']
    
    producto = db.execute('''
        SELECT * FROM productos
        WHERE id = %s AND usuario_id = %s
    ''', (id, user_id)).fetchone()
    
    db.close()
    
    if producto:
        return jsonify({
            'id': producto['id'],
            'nombre': producto['nombre'],
            'precio_venta': producto['precio_venta'],
            'costo_unitario': producto['costo_unitario'],
            'cantidad': producto['cantidad']
        })
    return jsonify({'error': 'Producto no encontrado'}), 404

# ==================== MAIN ====================

# Inicializar base de datos al importar
try:
    init_database()
    print("✓ Sistema inicializado correctamente")
except Exception as e:
    print(f"⚠️  Error al inicializar base de datos: {e}")
    print("   El sistema continuará, pero necesitas configurar DATABASE_URL")

if __name__ == '__main__':
    # Solo para desarrollo local
    debug_mode = os.environ.get('DEBUG', 'False').lower() == 'true'
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=debug_mode, host='0.0.0.0', port=port)
